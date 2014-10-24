# -*- coding: utf-8 -*-

import openpyxl
from gramgetter import GrammGetter as gget
ruvox=[]
ruvox.append("дом")
ruvox.append("ресторан")
ruvox.append("кофе")
ruvox.append("адвокат")
ruvox.append("флаг")
ruvox.append("школьник")
ruvox.append("Дания")
ruvox.append("барабан")
ruvox.append("думать")
ruvox.append("дышать")
ruvox.append("завтракать")
ruvox.append("значить")
ruvox.append("учить")
ruvox.append("приятный")
ruvox.append("какой")
ruvox.append("рациональный")
ruvox.append("жаркий")
ruvox.append("чистый")


wb= openpyxl.load_workbook("D:\\ppy_projekte\\russTrainer\\data\\Test1.xlsx")
ws=wb.get_sheet_by_name("Tabelle1")

for j,ru in enumerate(ruvox):
    info=gget.getWikiGrammar(ru)
##    #get total html text
##    htxt= gget.getHtmlText(ru)
##    print("got text")
##
##    #get category
##    cat=  gget.getCategory(htxt)
##    print ("got cat "+cat)
##
##    if cat!="o":
##        #get table
##        tabletext=gget.getTableText(htxt,cat)
##        print("laenge Table: "+str(len(tabletext)))
##
##        #get cells
##        cells=gget.getStringPart(tabletext,st="<td",en="/td>")
##
##    #call function
##    if cat == "a" :
##        cut_cells=gget.getWikiAdjectiv(ru,cells)
##    if cat == "n" :
##        cut_cells=gget.getWikiNomen(ru,cells)
##    if cat == "v" :
##        cut_cells=gget.getWikiVerb(ru,cells)
##    if cat == "o" :
##        cut_cells="--manually--"
##
##
    ws.cell(row=j+1,column=1).value=ru
    ws.cell(row=j+1,column=2).value=info
##    #ws.cell(row=j+1,column=3).value=cut_cells[0]

wb.save("D:\\ppy_projekte\\russTrainer\\data\\Test1.xlsx")