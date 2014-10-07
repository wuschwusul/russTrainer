# -*- coding: utf-8 -*-
import os
import openpyxl

from os.path import join

f_pt=os.path.realpath(__file__)
print f_pt

d_sub1= os.path.dirname(__file__)
print d_sub1

d_sub2 =os.path.dirname(d_sub1)
print d_sub2

d_sub3 =os.path.dirname(d_sub2)
print d_sub3

print os.path.dirname(os.path.dirname(os.path.dirname(__file__)))

##fn = "K:\ppy_projekte\prj_russTrainerSimpel\data\slawar.xlsx"
##
##wb= openpyxl.load_workbook(filename = "K:\ppy_projekte\prj_russTrainerSimpel\data\slawar.xlsx")
##ws=wb.get_sheet_by_name("Vocs")
####
##print ws.cell("H22").value
##
####
##for i,row in enumerate(ws.rows):
##    print row[0].value
##    print row[1].value
##    print row[2].value
##    print row[3].value
##    print row[4].value
##    print row[5].value
##    print row[6].value
##    print row[7].value
##    print row[8].value
##    if i==20:
##        break


##wb.save( "K:\ppy_projekte\prj_russTrainerSimpel\data\slawar.xlsx") # don't forget to save !
##
####    for cell in row:
####        print cell.value
##
##ws.cell(row=4,column=16).value="jesssas"
####
##wb.save(fn)


####
####de="wissen"
##sru="знать"
####ru=unicode(sru,'utf-8')
##
##brr=True
##
##print "hui"+unicode(sru,'utf-8')

#import time
##from time import strftime
##
##print (strftime("%Y%m%d_%H%M%S"))


from os import listdir
##
##for f in listdir("K:\\ppy_projekte\\prj_russTrainerSimpel\\data\\archiv"):
##    print f
##
##print("xxx")
##print listdir("K:\\ppy_projekte\\prj_russTrainerSimpel\\data\\archiv")[-1]
##
##print listdir("K:\\test")
##print len(listdir("K:\\test"))

##
##
##
##db_dir="K:\\ppy_projekte\\prj_russTrainerSimpel\\data"
##print listdir(join(db_dir,"archiv"))

##from gramgetter import GrammGetter as gget
##
##print("startin....")
##ru="знать"
##tt= gget.getWikiText(ru)
##
##for line in tt:
##    print line
def getStringPart(text,st,en):

    #tt="<td bgcolor=kneydlwoika</td>zwischen<td bgcolor=kneydlwoika</td>auusss"
    fin=[]

    for l in text.split(en):
        if st in l:
            fin.append(l.split(st)[1])

    return fin



print("Wikipedia HTML Testsuchlauf")

ftxt= open("K:\\ppy_projekte\\prj_russTrainerSimpel\\app\\testhtmls\\dom.htm")
#ftxt= open("K:\\ppy_projekte\\prj_russTrainerSimpel\\app\\testhtmls\\malenki.htm")
#ftxt= open("K:\\ppy_projekte\\prj_russTrainerSimpel\\app\\testhtmls\\snat.htm")


print("START Category"),
# KATEGORIESUCHE
        #signalwoerter
        #nomen=Существительное
        #verb=Глагол
        #adj = Прилагательное
found=False
for i,line in enumerate(ftxt):
    if "Существительное" in line:
        print("FOUND-is a nomen")
        found=True
        break
    elif "Глагол" in line:
        print ("FOUND-is a verb")
        found=True
        break
    elif "Прилагательное" in line:
        print ("FOUND-is a adjectiv")
        found=True
        break

if found==False:
    print("NOT FOUND-it is other")


ftxt= open("K:\\ppy_projekte\\prj_russTrainerSimpel\\app\\testhtmls\\dom.htm")
#GRAMMATIKTAbllenSuche
print("START search Grammartable"),
tabletext=""
wiki_record=0

for i,line in enumerate(ftxt):
    #search start
    if "падеж</a>" in line:
        print("FOUND - start record at line "+str(i))
        wiki_record=1

    #record
    if(wiki_record==1):
        tabletext+=line
    #search end
    if ('</table' in line) and (wiki_record==1):
        wiki_record=0
        print("END recording")

print("found table")
#print tabletext


cellz= getStringPart(tabletext,st="<td",en="/td>")

print cellz[4]

##print("number rows")
##print len(tabletext.split("</tr")) #splitten der reihen
##
##for rw in tabletext.split("</tr"):
##    print("NEEEEEEEEEEXTROW")
##    print rw




##for i,line in enumerate(ftxt):
##    if i==127:
##        print line
##        print line.decode("utf-8")
##        if "падеж" in line:
##            print("found")

##
##print("...start search")
##for i,line in enumerate(ftxt):
##    if
##        print("found at"+str(i))



