# -*- coding: utf-8 -*-
from gramgetter import GrammGetter as gget

from vocabel import Vocabel
from time import strftime
from os.path import join,dirname
from os import listdir
import sys

##import xlrd
##import xlwt

sys.path.append(join(dirname(dirname(__file__)),"lib"))
sys.path.append(join(dirname(dirname(__file__)),"lib","openpyxl"))

import openpyxl

class Slawar:
    vocs=None
    vocmax=0  #geladene Anzahl an vocs, neue kommen danach
    wb=None
    ws=None
    prj_dir=""
    db_dir=""               #"K:\\ppy_projekte\\prj_russTrainerSimpel\\data"
    db_filename=""              #"K:\\ppy_projekte\\prj_russTrainerSimpel\\data\\slawar.xlsx"
                                    ##db_filename2=join(db_dir,"slawar2.xlsx")


    def __init__(self):
        print(">>START initialize Paths and Objects..."),

        self.vocs=[]
        self.prj_dir= dirname(dirname(dirname(__file__)))    #project directory
        self.db_dir= join(self.prj_dir,"data")                                            #data directory
        self.db_filename=join(self.db_dir,"slawar.xlsx")

        print("...OK")



    def save(self, isArch=False):  #saving Workbook in Memory to Databasefile or Creates new ArchivFile with timestamp

        print(">>START saving Data ..."),
        if len(self.vocs)==0:
            print("ERROR - No Database loaded, saving aborted -XXX")
            return False

        offset=2 # row-offset in excel
        for i,rw in enumerate(self.vocs):
            self.ws.cell(row=i+offset,column=1).value=i
            self.ws.cell(row=i+offset,column=2).value=self.vocs[i].ru
            self.ws.cell(row=i+offset,column=3).value=self.vocs[i].de
            self.ws.cell(row=i+offset,column=4).value=self.vocs[i].cat
            self.ws.cell(row=i+offset,column=5).value=self.vocs[i].gmr
            self.ws.cell(row=i+offset,column=8).value=self.vocs[i].ktx

        if isArch==False:
            #self.wb.save(self.db_filename)
            self.wb.save(self.db_filename)
            print("...to Database ...OK")

        if isArch==True:                            # saving as an ARchiv
            ts=(strftime("%Y%m%d_%H%M%S"))          # timestamp
            self.wb.save(join(self.db_dir,"archiv","ar_"+ts+".xlsx"))
            print("...to Archiv ...OK")



    def load(self, isArch=False):
        print(">>START loading Data..."),

        if isArch==False: #then load standard database
            print("...from Database..."),
            fileToLoad=self.db_filename

        if isArch==True:     #then load latest archiv database
            print("...from recent Archiv..."),
            fileToLoad=self.getLastArchivFileName()

            if fileToLoad==False:
                print("FILE OPEN ERROR - no file found...XXX")
                return False

        self.loadDBsheet(fileToLoad,"Vocs")


        print("...fillin Vocs..."),
        del self.vocs[:]  #clear vocs array

        for rw,rowfill in enumerate(self.ws.rows):                          #iterate through rows
            if rw!=0:
                nr=rowfill[0].value  # col=0->Number of voc
                ru=rowfill[1].value  # col=1->ru
                de=rowfill[2].value  # col=2->de
                ct=rowfill[3].value  # col=3->cat
                gm=rowfill[4].value  # col=4>gmr
                kx=rowfill[7].value  # col=7->kontext

                #fill class variable
                self.vocmax+=1
                self.add(de,ru,ct,gm,kx)

        print("OK")


    def loadDBsheet(self,f,s):  # loads sheet to class attribute "ws" #f=filename, s=sheetname
        self.wb= openpyxl.load_workbook(filename=f,data_only=True)
        self.ws=self.wb.get_sheet_by_name(s)



    def show(self):
        print("show all entries")
        ShowRus=True

        #check if cyrillic can be displayed (eg. in MS DOS not, in PYthon console it can)
        try:
            print("Check: "+v.ru)
        except:
            print("cyrillic symbols cannot be displayed - only german vocs are shown")
            ShowRus=False

        for i,v in enumerate(self.vocs):
            if ShowRus==False:
                print("Nr."+str(i)+":"+v.de+","+v.cat+","+v.ktx) # ohne russ.buchst
            else:
                print("Nr."+str(i)+":"+v.de+","+v.ru+","+v.cat+","+v.ktx)



    def fillEmtpyGram(self):
        print(">>START filling Empty Grammar Cells...")
        cnt=0
        print(">>COUNTING...")              # count the empty cells
        for i,v in enumerate(self.vocs):
            if v.gmr==None:
                cnt+=1
        if cnt==0:
            print("All entries already filled - nothing added")
            return False

        # for every empty, get them
        for i,v in enumerate(self.vocs):
            if v.gmr==None:
                v.gmr= gget.getWikiGrammar(v.ru,i,cnt)









    def getLastArchivFileName(self): #!! returns last archiv filename
        arpath=join(self.db_dir,"archiv")   #archiv directory
        flz=listdir(arpath)                 #get all filez in archiv

        for i,entry in enumerate(flz):  #search for a file with "ar_.." from the bottom of thelist =latest
            if flz[-i-1][:3]=="ar_":
                af=flz[-i-1]
                print("useing archiv "+af),
                return join(arpath,af)

        print("LOADING ERROR - No Archives found -XXX")
        return False #no archiv existing


    def isEmpty(self):
        if len(self.vocs)==0:
            return True
        return False


    def add(self,de,ru,ct="",gm="",kx=""):  #("word","слова","n")
        #print("adding word, DEACTIVATED")
        self.vocs.append(Vocabel(de,ru,cat=ct,ktx=kx,gmr=gm))

    def delete(self,de):
        print("deleting entry: "+de)
        print("not implemented yet")



########### MP3 related ###########
    def getMp3(self):
        pass

    def getMp3all(self):
        pass

    def getMp3Missing(self):
        pass

